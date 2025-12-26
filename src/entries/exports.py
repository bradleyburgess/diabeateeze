"""Export functionality for entries models."""

import csv
import io
from datetime import datetime
from typing import Any, List

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import GlucoseReading, InsulinDose, Meal


class BaseExporter:
    """Base class for data exporters."""

    def __init__(self, queryset, model_class):
        self.queryset = queryset
        self.model_class = model_class

    def get_headers(self) -> List[str]:
        """Return column headers for export."""
        raise NotImplementedError

    def get_row_data(self, obj: Any) -> List[Any]:
        """Return row data for a single object."""
        raise NotImplementedError

    def to_csv(self) -> HttpResponse:
        """Export data to CSV format."""
        response = HttpResponse(content_type="text/csv")
        filename = f"{self.model_class.__name__.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(self.get_headers())

        for obj in self.queryset:
            writer.writerow(self.get_row_data(obj))

        return response

    def to_excel(self) -> HttpResponse:
        """Export data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = self.model_class.__name__

        # Add headers with styling
        headers = self.get_headers()
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Add data rows
        for row_idx, obj in enumerate(self.queryset, start=2):
            for col_idx, value in enumerate(self.get_row_data(obj), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"{self.model_class.__name__.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    def to_text(self) -> HttpResponse:
        """Export data to plain text format."""
        response = HttpResponse(content_type="text/plain; charset=utf-8")
        filename = f"{self.model_class.__name__.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        lines = []
        for obj in self.queryset:
            lines.append(self.format_text_entry(obj))
            lines.append("")  # Empty line between entries

        response.write("\n".join(lines))
        return response

    def format_text_entry(self, obj: Any) -> str:
        """Format a single entry as text. Override in subclasses."""
        return str(obj)


class GlucoseReadingExporter(BaseExporter):
    """Exporter for GlucoseReading model."""

    def get_headers(self) -> List[str]:
        return ["Date", "Time", "Value", "Unit", "Notes"]

    def get_row_data(self, obj: GlucoseReading) -> List[Any]:
        return [
            obj.occurred_at.strftime("%Y-%m-%d"),
            obj.occurred_at.strftime("%H:%M:%S"),
            float(obj.value),
            obj.unit,
            obj.notes,
        ]

    def to_text(self) -> HttpResponse:
        """Export glucose readings as plain text with bulleted list format."""
        response = HttpResponse(content_type="text/plain; charset=utf-8")
        filename = f"{self.model_class.__name__.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        lines = []

        # Determine date range from queryset
        readings_list = list(self.queryset)
        if readings_list:
            dates = [r.occurred_at for r in readings_list]
            earliest = min(dates)
            latest = max(dates)

            if earliest.date() == latest.date():
                date_range = earliest.strftime("%Y/%m/%d")
            else:
                date_range = (
                    f"{earliest.strftime('%Y/%m/%d')} - {latest.strftime('%Y/%m/%d')}"
                )
        else:
            date_range = "No Data"

        lines.append(f"Glucose Readings for {date_range}")

        for reading in readings_list:
            lines.append(self.format_text_entry(reading))

        response.write("\n".join(lines))
        return response

    def format_text_entry(self, obj: GlucoseReading) -> str:
        time_str = (
            obj.occurred_at.strftime("%Y/%m/%d %I:%M %p").replace(" 0", " ").lower()
        )
        return f"- {time_str}: {obj.value} {obj.unit}"


class InsulinDoseExporter(BaseExporter):
    """Exporter for InsulinDose model."""

    def get_headers(self) -> List[str]:
        return [
            "Date",
            "Time",
            "Insulin Type",
            "Base Units",
            "Correction Units",
            "Total Units",
            "Notes",
        ]

    def get_row_data(self, obj: InsulinDose) -> List[Any]:
        total = float(obj.base_units + obj.correction_units)
        return [
            obj.occurred_at.strftime("%Y-%m-%d"),
            obj.occurred_at.strftime("%H:%M:%S"),
            obj.insulin_type.name,
            float(obj.base_units),
            float(obj.correction_units),
            total,
            obj.notes or "",
        ]

    def format_text_entry(self, obj: InsulinDose) -> str:
        total = obj.base_units + obj.correction_units
        text = f"Insulin Dose - {obj.occurred_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
        text += f"  Type: {obj.insulin_type.name}\n"
        text += f"  Base: {obj.base_units} units\n"
        text += f"  Correction: {obj.correction_units} units\n"
        text += f"  Total: {total} units"
        if obj.notes:
            text += f"\n  Notes: {obj.notes}"
        return text


class MealExporter(BaseExporter):
    """Exporter for Meal model."""

    def get_headers(self) -> List[str]:
        return ["Date", "Time", "Meal Type", "Description", "Total Carbs (g)", "Notes"]

    def get_row_data(self, obj: Meal) -> List[Any]:
        return [
            obj.occurred_at.strftime("%Y-%m-%d"),
            obj.occurred_at.strftime("%H:%M:%S"),
            obj.get_meal_type_display(),
            obj.description,
            float(obj.total_carbs) if obj.total_carbs else "",
            obj.notes,
        ]

    def format_text_entry(self, obj: Meal) -> str:
        text = f"{obj.get_meal_type_display()} - {obj.occurred_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
        text += f"  Description: {obj.description}"
        if obj.total_carbs:
            text += f"\n  Carbs: {obj.total_carbs}g"
        if obj.notes:
            text += f"\n  Notes: {obj.notes}"
        return text


def get_exporter_for_model(model_class):
    """Factory function to get the appropriate exporter for a model."""
    exporters = {
        GlucoseReading: GlucoseReadingExporter,
        InsulinDose: InsulinDoseExporter,
        Meal: MealExporter,
    }
    return exporters.get(model_class, BaseExporter)
